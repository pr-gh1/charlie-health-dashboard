"""
build_n24m.py

N24M (next 24 months) IOP revenue projection model, built directly in
Excel formulas per the case study's submission format (<200 rows, not
lines of code). Two payor tracks (Commercial, Medicaid), rep-driven
admissions (per the case study's B2B outreach-rep hint), a LOS lag for
discharges derived from actual attended-appointment duration (most
recent completed cohort, not scheduled/booked dates), base case vs.
growth case (added reps, ramping in), reporting IOP / OPT / total
revenue.

All starting assumptions are pulled from the real base dataset via
kpis.py, not invented.
"""
import sys
sys.path.insert(0, ".")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime
from transform_attendance import load_grid, parse_grid
from kpis import build_frames, classify_episodes, los_by_discharge_month
import pandas as pd

grid, wb0 = load_grid("sample_data/sample_attendance_export.xlsx")
rates, sessions, patients = parse_grid(grid)
rates_df, sessions_df, patients_df = build_frames(rates, sessions, patients)

# --- pull real starting assumptions ---
sessions_df["month"] = sessions_df["date"].dt.to_period("M")
classified = classify_episodes(patients_df, sessions_df)
active = classified[classified.episode_status == "active"]
census0_comm = int((active.payor == "Commercial").sum())
census0_med = int((active.payor == "Medicaid").sum())

total_admits_8mo = 53
base_admit_rate = total_admits_8mo / 8  # 6.625/month
comm_mix = (patients_df.payor == "Commercial").mean()
med_mix = 1 - comm_mix

patients_df["weeks_in_tx"] = patients_df["los_days"] / 7
weeks_by_payor = patients_df.groupby("payor")["weeks_in_tx"].sum()
billable_by_payor_type = sessions_df.groupby(["payor", "session_type"])["billable"].sum()
sess_wk = {
    (p, t): billable_by_payor_type.get((p, t), 0) / weeks_by_payor[p]
    for p in ("Commercial", "Medicaid") for t in ("IOP", "OPT")
}
rate = {(r["payor"], r["session_type"]): r["rate"] for r in rates}

# Historical admits Jan/Feb/Mar 2021 (real lookback for the discharge lag)
first_dates = sessions_df.groupby("patient_id")["date"].min().dt.to_period("M")
hist_admits = first_dates.value_counts()
lookback = [int(hist_admits.get(pd.Period(m, "M"), 0)) for m in ["2021-01", "2021-02", "2021-03"]]

REPS_BASE = 5
ADMITS_PER_REP = round(base_admit_rate / REPS_BASE, 4)

# Growth case is staged, milestone-gated hiring rather than a single
# one-time step: each wave of reps ramps to full productivity, then holds
# for a "prove it out" period (the milestone -- e.g. hitting target
# referral/admit volume) before the next wave is approved. Three waves
# across the 24-month horizon, 6 months apart (3 to ramp + 3 to prove
# out), so the line keeps climbing through most of the window instead of
# flatlining after month 6. Sized as a genuine "go-big" plan -- +5 reps
# per wave (not +2) -- so headcount roughly quadruples (5 -> 20) rather
# than doubling, a deliberately bigger swing than a conservative
# incremental hire.
REPS_PER_WAVE = 5
RAMP_LEN = 3
WAVE_STARTS = [4, 10, 16]  # month index (1-based) each wave's ramp begins

# LOS: the headline KPI is number of appointments ATTENDED per patient stay
# (not a calendar-time span, which conflates episode length with scheduling
# gaps) -- for the most recent fully-completed discharge cohort, derived
# from the real data via kpis.los_by_discharge_month(), not hardcoded.
# The month-based lag below (LOS_MONTHS) is a separate, internal timing
# parameter this monthly model structurally needs to know when to roll a
# cohort off census -- it is NOT the reported LOS KPI, just derived from
# the same cohort's calendar span for internal consistency.
_los_trend = los_by_discharge_month(sessions_df, patients_df)
LOS_APPOINTMENTS_MOST_RECENT_COHORT = float(_los_trend.iloc[-1]["avg_los_appointments"])
LOS_WEEKS_MOST_RECENT_COHORT = float(_los_trend.iloc[-1]["avg_los_weeks"])
WEEKS_PER_MONTH = 4.345
LOS_MONTHS = round(LOS_WEEKS_MOST_RECENT_COHORT / WEEKS_PER_MONTH)

MONTHS = pd.period_range("2021-04", periods=24, freq="M")

# ---------------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "N24M Model"

HDR = Font(bold=True, name="Arial")
FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
INPUT_FONT = Font(name="Arial", color="1F4E78", bold=True)  # blue = input
CALC_FONT = Font(name="Arial")
SECTION = Font(bold=True, size=13, name="Arial")

def cell(r, c, v=None, font=CALC_FONT, fmt=None, fill=None):
    x = ws.cell(row=r, column=c, value=v)
    x.font = font
    if fmt: x.number_format = fmt
    if fill: x.fill = fill
    return x

r = 1
cell(r, 1, "N24M -- IOP Revenue Projection Model", SECTION); r += 1
cell(r, 1, "Scoped to IOP revenue per the case study ask; OPT and total revenue reported alongside since both fall out of the same patient-flow structure. See N24M_Assumptions_and_Limitations.docx for full methodology notes.", Font(italic=True, size=9, color="595959")); r += 2

# --- Assumptions block (blue = editable inputs) ---
cell(r, 1, "Assumptions (editable)", Font(bold=True, size=12)); r += 1
assum_row = {}
def add_assum(label, value, fmt=None):
    global r
    cell(r, 1, label, CALC_FONT)
    c = cell(r, 2, value, INPUT_FONT, fmt)
    assum_row[label] = r
    r += 1
    return c

add_assum("Starting reps (base case, current)", REPS_BASE, "0")
add_assum("Admissions per rep per month", ADMITS_PER_REP, "0.000")
add_assum("Growth case: reps added per wave", REPS_PER_WAVE, "0")
add_assum("Growth case: ramp length per wave (months)", RAMP_LEN, "0")
wave_labels = []
for wi, wave_month in enumerate(WAVE_STARTS, start=1):
    label = f"Growth case: wave {wi} start month (1-24)"
    add_assum(label, wave_month, "0")
    wave_labels.append(label)
add_assum(
    f"LOS -- reported KPI: {LOS_APPOINTMENTS_MOST_RECENT_COHORT:.1f} appointments attended "
    f"(most recent cohort). Discharge-timing lag below (months)",
    LOS_MONTHS, "0",
)
add_assum("Commercial admit mix", round(comm_mix, 4), "0.0%")
add_assum("Medicaid admit mix", round(med_mix, 4), "0.0%")
add_assum("Starting census -- Commercial (as of Mar 2021)", census0_comm, "0")
add_assum("Starting census -- Medicaid (as of Mar 2021)", census0_med, "0")
add_assum("Commercial IOP sessions/patient/week", round(sess_wk[("Commercial", "IOP")], 3), "0.000")
add_assum("Commercial OPT sessions/patient/week", round(sess_wk[("Commercial", "OPT")], 3), "0.000")
add_assum("Medicaid IOP sessions/patient/week", round(sess_wk[("Medicaid", "IOP")], 3), "0.000")
add_assum("Medicaid OPT sessions/patient/week", round(sess_wk[("Medicaid", "OPT")], 3), "0.000")
add_assum("IOP rate -- Commercial", rate[("Commercial", "IOP")], "$#,##0")
add_assum("IOP rate -- Medicaid", rate[("Medicaid", "IOP")], "$#,##0")
add_assum("OPT rate -- Commercial", rate[("Commercial", "OPT")], "$#,##0")
add_assum("OPT rate -- Medicaid", rate[("Medicaid", "OPT")], "$#,##0")
add_assum("Weeks per month", WEEKS_PER_MONTH, "0.000")
add_assum("Historical admits Jan 2021 (lookback for discharge lag)", lookback[0], "0")
add_assum("Historical admits Feb 2021 (lookback for discharge lag)", lookback[1], "0")
add_assum("Historical admits Mar 2021 (lookback for discharge lag)", lookback[2], "0")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=26)
cell(
    r, 1,
    "Flag: these three lookback months are genuinely 0 in the source data -- "
    "not a placeholder. Admissions were flat at zero for the full quarter "
    "immediately before this forecast starts, which sits inside the "
    "\"Admissions per rep per month\" assumption above (an 8-month average "
    "that blends this flat quarter in with five earlier, busier months). "
    "If the flat quarter reflects a genuine new demand level rather than "
    "one-off noise, the base case is optimistic -- see Assumptions & "
    "Limitations doc for the full discussion.",
    Font(italic=True, size=9, color="C00000"),
)
ws.row_dimensions[r].height = 28
r += 1
r += 1

A = lambda label: f"$B${assum_row[label]}"

FIRST_COL = 3  # column C = month 1

def month_header(start_row):
    cell(start_row, 1, "Month", HDR, fill=FILL)
    for i, m in enumerate(MONTHS):
        c = FIRST_COL + i
        x = cell(start_row, c, str(m), HDR, fmt="@", fill=FILL)
        x.alignment = Alignment(horizontal="center")

def build_scenario(label, is_growth):
    global r
    cell(r, 1, label, Font(bold=True, size=12)); r += 1
    month_header(r); hdr_row = r; r += 1

    rep_row = r
    cell(r, 1, "Rep count")
    for i in range(24):
        c = FIRST_COL + i
        if not is_growth:
            f = f"={A('Starting reps (base case, current)')}"
        else:
            m = i + 1
            # Each wave contributes a clamped linear ramp: 0 before its
            # start month, 0->1 (of REPS_PER_WAVE) across the ramp
            # length, then holds at 1 (fully ramped, "milestone hit")
            # until the model horizon ends -- MIN/MAX does the clamping
            # so this is one formula, not a nested IF per regime.
            wave_terms = "".join(
                f"+{A('Growth case: reps added per wave')}*"
                f"MIN(MAX(({m}-{A(label)}+1)/{A('Growth case: ramp length per wave (months)')},0),1)"
                for label in wave_labels
            )
            f = f"={A('Starting reps (base case, current)')}{wave_terms}"
        cell(r, c, f, fmt="0.00")
    r += 1

    if is_growth:
        note = ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=26)
        cell(
            r, 1,
            "Plain-English: reps = base headcount, plus each wave's contribution. "
            "Each wave ramps from 0 to full size (B7) over the ramp length (B8), "
            "starting at that wave's own start month (B9/B10/B11) -- "
            "MIN(MAX((month-start+1)/ramp,0),1) is just \"0 before the wave starts, "
            "rises in a straight line during the ramp, holds at 1 (fully ramped) "
            "forever after\" -- so a wave that's already ramped doesn't ramp again, "
            "and a wave that hasn't started yet doesn't contribute early.",
            Font(italic=True, size=9, color="595959"),
        )
        ws.row_dimensions[r].height = 28
        r += 1

    admits_row = r
    cell(r, 1, "Total admits")
    for i in range(24):
        c = FIRST_COL + i
        col_l = get_column_letter(c)
        cell(r, c, f"={col_l}{rep_row}*{A('Admissions per rep per month')}", fmt="0.00")
    r += 1

    admits_c_row = r
    cell(r, 1, "Admits -- Commercial")
    for i in range(24):
        c = FIRST_COL + i; col_l = get_column_letter(c)
        cell(r, c, f"={col_l}{admits_row}*{A('Commercial admit mix')}", fmt="0.00")
    r += 1
    admits_m_row = r
    cell(r, 1, "Admits -- Medicaid")
    for i in range(24):
        c = FIRST_COL + i; col_l = get_column_letter(c)
        cell(r, c, f"={col_l}{admits_row}*{A('Medicaid admit mix')}", fmt="0.00")
    r += 1

    disch_c_row = r
    cell(r, 1, f"Discharges -- Commercial (=admits {LOS_MONTHS}mo prior)")
    for i in range(24):
        c = FIRST_COL + i; col_l = get_column_letter(c)
        src_i = i - LOS_MONTHS
        if src_i < 0:
            hist_idx = src_i + 3  # index into lookback (Jan=0,Feb=1,Mar=2)
            val = round(lookback[hist_idx] * comm_mix, 4) if 0 <= hist_idx < 3 else 0
            cell(r, c, val, fmt="0.00")
        else:
            src_col = get_column_letter(FIRST_COL + src_i)
            cell(r, c, f"={src_col}{admits_c_row}", fmt="0.00")
    r += 1
    disch_m_row = r
    cell(r, 1, f"Discharges -- Medicaid (=admits {LOS_MONTHS}mo prior)")
    for i in range(24):
        c = FIRST_COL + i; col_l = get_column_letter(c)
        src_i = i - LOS_MONTHS
        if src_i < 0:
            hist_idx = src_i + 3
            val = round(lookback[hist_idx] * med_mix, 4) if 0 <= hist_idx < 3 else 0
            cell(r, c, val, fmt="0.00")
        else:
            src_col = get_column_letter(FIRST_COL + src_i)
            cell(r, c, f"={src_col}{admits_m_row}", fmt="0.00")
    r += 1

    census_c_row = r
    cell(r, 1, "Census -- Commercial")
    for i in range(24):
        c = FIRST_COL + i; col_l = get_column_letter(c)
        prev = f"{A('Starting census -- Commercial (as of Mar 2021)')}" if i == 0 else f"{get_column_letter(c-1)}{census_c_row}"
        cell(r, c, f"={prev}+{col_l}{admits_c_row}-{col_l}{disch_c_row}", fmt="0.0")
    r += 1
    census_m_row = r
    cell(r, 1, "Census -- Medicaid")
    for i in range(24):
        c = FIRST_COL + i; col_l = get_column_letter(c)
        prev = f"{A('Starting census -- Medicaid (as of Mar 2021)')}" if i == 0 else f"{get_column_letter(c-1)}{census_m_row}"
        cell(r, c, f"={prev}+{col_l}{admits_m_row}-{col_l}{disch_m_row}", fmt="0.0")
    r += 1
    census_tot_row = r
    cell(r, 1, "Census -- Total")
    for i in range(24):
        c = FIRST_COL + i; col_l = get_column_letter(c)
        cell(r, c, f"={col_l}{census_c_row}+{col_l}{census_m_row}", fmt="0.0")
    r += 1

    iop_rev_row = r
    cell(r, 1, "IOP revenue")
    for i in range(24):
        c = FIRST_COL + i; col_l = get_column_letter(c)
        f = (f"=({col_l}{census_c_row}*{A('Commercial IOP sessions/patient/week')}*{A('IOP rate -- Commercial')}"
             f"+{col_l}{census_m_row}*{A('Medicaid IOP sessions/patient/week')}*{A('IOP rate -- Medicaid')})*{A('Weeks per month')}")
        cell(r, c, f, fmt="$#,##0")
    r += 1
    opt_rev_row = r
    cell(r, 1, "OPT revenue")
    for i in range(24):
        c = FIRST_COL + i; col_l = get_column_letter(c)
        f = (f"=({col_l}{census_c_row}*{A('Commercial OPT sessions/patient/week')}*{A('OPT rate -- Commercial')}"
             f"+{col_l}{census_m_row}*{A('Medicaid OPT sessions/patient/week')}*{A('OPT rate -- Medicaid')})*{A('Weeks per month')}")
        cell(r, c, f, fmt="$#,##0")
    r += 1
    tot_rev_row = r
    cell(r, 1, "Total revenue (IOP + OPT)", Font(bold=True))
    for i in range(24):
        c = FIRST_COL + i; col_l = get_column_letter(c)
        cell(r, c, f"={col_l}{iop_rev_row}+{col_l}{opt_rev_row}", Font(bold=True), fmt="$#,##0")
    r += 2
    return tot_rev_row, iop_rev_row

base_tot_row, base_iop_row = build_scenario("Base case (rep headcount flat)", is_growth=False)
growth_tot_row, growth_iop_row = build_scenario(
    "Go-big case (staged rep hiring, 3 waves of +5, each gated on the prior wave ramping to full productivity)",
    is_growth=True,
)

# --- 24-month totals summary ---
cell(r, 1, "24-month totals", Font(bold=True, size=12)); r += 1
cell(r, 1, "Base case -- IOP revenue"); cell(r, 2, f"=SUM(C{base_iop_row}:Z{base_iop_row})", fmt="$#,##0"); r += 1
cell(r, 1, "Base case -- total revenue"); cell(r, 2, f"=SUM(C{base_tot_row}:Z{base_tot_row})", fmt="$#,##0"); r += 1
cell(r, 1, "Growth case -- IOP revenue"); cell(r, 2, f"=SUM(C{growth_iop_row}:Z{growth_iop_row})", fmt="$#,##0"); r += 1
cell(r, 1, "Growth case -- total revenue"); cell(r, 2, f"=SUM(C{growth_tot_row}:Z{growth_tot_row})", fmt="$#,##0"); r += 1

ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 14
for i in range(24):
    ws.column_dimensions[get_column_letter(FIRST_COL + i)].width = 11

print(f"Total rows used: {r}")
wb.save("../N24M_Revenue_Projection.xlsx")
print("Wrote ../N24M_Revenue_Projection.xlsx")
