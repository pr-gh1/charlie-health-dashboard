"""
extract_n24m_baseline.py

Reads the recalculated N24M_Revenue_Projection.xlsx (values, not formulas)
and writes a frozen forecast baseline JSON the dashboard can read directly
-- this is the "generate once, compare against forever after" artifact
from the toggle-bar/forecast architecture design.

Row layout is located dynamically by label in column A, not hardcoded, so
this survives minor row-count changes to the model (same resilience
principle as transform_attendance.py's label-driven parsing).
"""
import json
import sys
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "/tmp/recalced2/N24M_Revenue_Projection.xlsx"
OUT = "forecast/n24m_baseline.json"

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["N24M Model"]

def find_row(label_substr, start=1):
    for r in range(start, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and label_substr.lower() in str(v).lower():
            return r
    raise ValueError(f"Row not found: {label_substr}")

def read_row(r, ncols=24, first_col=3):
    return [ws.cell(row=r, column=first_col + i).value for i in range(ncols)]

base_start = find_row("Base case (rep headcount")
growth_start = find_row("Growth case (+")

month_row = find_row("Month", base_start)
months = [str(m) for m in read_row(month_row)]

def scenario_block(start_row):
    census_row = find_row("Census -- Total", start_row)
    admits_row = find_row("Total admits", start_row)
    iop_row = find_row("IOP revenue", start_row)
    opt_row = find_row("OPT revenue", start_row)
    tot_row = find_row("Total revenue", start_row)
    return {
        "census": [round(v, 1) if v is not None else None for v in read_row(census_row)],
        "admits": [round(v, 1) if v is not None else None for v in read_row(admits_row)],
        "iop_revenue": [round(v, 2) if v is not None else None for v in read_row(iop_row)],
        "opt_revenue": [round(v, 2) if v is not None else None for v in read_row(opt_row)],
        "total_revenue": [round(v, 2) if v is not None else None for v in read_row(tot_row)],
    }

baseline = {
    "generated_from": "N24M_Revenue_Projection.xlsx",
    "baseline_cutoff": "2021-03-31",
    "months": months,
    "base_case": scenario_block(base_start),
    "growth_case": scenario_block(growth_start),
}

import os
os.makedirs("forecast", exist_ok=True)
with open(OUT, "w") as f:
    json.dump(baseline, f, indent=2)

print(f"Wrote {OUT}")
print(f"Months: {months[0]} to {months[-1]} ({len(months)} months)")
print(f"Base case total_revenue[0:3]: {baseline['base_case']['total_revenue'][:3]}")
print(f"Growth case total_revenue[0:3]: {baseline['growth_case']['total_revenue'][:3]}")
