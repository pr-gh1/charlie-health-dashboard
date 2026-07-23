"""
merge_workbook.py

Consolidates all three case study parts into a single workbook, built on
top of the user's master file (Charlie_Health_Case_Study.xlsx), which
already carries the two raw source tabs (Attendance and Billing, Acme
Dash) plus the Part 2 model. This adds the Part 1 transform/model tabs
and the Part 3 backup-analysis tabs, so the whole submission is one file
instead of four.

Cross-workbook sheet copy isn't natively supported by openpyxl, so this
copies cell-by-cell (value + style + number format), plus column widths
and merged-cell ranges, then reorders the final sheet list.
"""
import openpyxl
from copy import copy

MASTER = "Charlie_Health_Case_Study.xlsx"
OUT = "Charlie_Health_Case_Study_Combined.xlsx"

master = openpyxl.load_workbook(MASTER)  # keep formulas (not data_only)


def copy_sheet(src_ws, title):
    dst_ws = master.create_sheet(title)
    for row in src_ws.iter_rows():
        for c in row:
            new_cell = dst_ws.cell(row=c.row, column=c.column, value=c.value)
            if c.has_style:
                new_cell.font = copy(c.font)
                new_cell.border = copy(c.border)
                new_cell.fill = copy(c.fill)
                new_cell.number_format = c.number_format
                new_cell.protection = copy(c.protection)
                new_cell.alignment = copy(c.alignment)
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[col_letter].width = dim.width
    for row_idx, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst_ws.row_dimensions[row_idx].height = dim.height
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))
    dst_ws.freeze_panes = src_ws.freeze_panes
    return dst_ws


# --- Part 1: transformed data + KPI summary + N24M model ---
kpi_wb = openpyxl.load_workbook("Charlie_Health_KPI_Summary_and_Sessions.xlsx")
for name in ["Rates", "Sessions", "Patients", "KPI Summary"]:
    copy_sheet(kpi_wb[name], name)
    print(f"Copied Part 1 tab: {name}")

n24m_wb = openpyxl.load_workbook("N24M_Revenue_Projection.xlsx")
copy_sheet(n24m_wb["N24M Model"], "N24M Model")
print("Copied Part 1 tab: N24M Model")

# --- Part 3: backup analysis (raw tab skipped -- master already has the
# same raw data as "Acme Dash"). Rep Analysis / Summary formulas reference
# it by its original name, "Acme Dash (Raw)", so the master's tab is
# renamed to match rather than duplicating the raw data under a second
# tab name -- nothing else in the master workbook references "Acme Dash"
# by formula (checked directly), so this rename is safe.
master["Acme Dash"].title = "Acme Dash (Raw)"

backup_wb = openpyxl.load_workbook("Acme_Dash_Backup_Analysis.xlsx")
for name in ["Rep Analysis", "Summary"]:
    copy_sheet(backup_wb[name], name)
    print(f"Copied Part 3 tab: {name}")

# --- Final tab order: Part 1 (raw -> transform -> model), Part 2, Part 3 (raw -> analysis) ---
desired_order = [
    "Attendance and Billing", "Rates", "Sessions", "Patients", "KPI Summary", "N24M Model",
    "Part 2 - New Products",
    "Acme Dash (Raw)", "Rep Analysis", "Summary",
]
missing = set(desired_order) - set(master.sheetnames)
extra = set(master.sheetnames) - set(desired_order)
assert not missing, f"Missing sheets: {missing}"
assert not extra, f"Unexpected extra sheets: {extra}"
master._sheets = [master[name] for name in desired_order]

master.save(OUT)
print(f"\nWrote {OUT}")
print("Final sheet order:", master.sheetnames)
