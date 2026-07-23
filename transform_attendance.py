"""
transform_attendance.py

Repeatable ETL pipeline for Charlie Health's Attendance & Billing export.

WHAT IT DOES
------------
The raw export is a "wide" grid: one column per calendar date, and each
patient occupies two rows (an IOP row and an OPT row stacked directly below
it, with the payor name tucked under the patient's name in column A). That
layout is fine for a clinician marking attendance by hand, but it can't be
grouped, filtered, or trended by patient / payor / session type / date --
which is exactly what the downstream KPI and revenue work needs.

This script unpivots (melts) the raw grid into a long "Sessions" table --
one row per *scheduled* session (blank cells, i.e. days nothing was
scheduled, are dropped) -- plus a Patients dimension table and a Rates
lookup table. Those three tables are the single source of truth that every
KPI, pivot, and the revenue projection model should build from.

WHY IT'S BUILT THIS WAY
------------------------
This is meant to run every time a new export arrives, not just once by
hand. So:
  - Parsing is driven by *labels* found in the sheet ("Patient info",
    "Rates"), not hardcoded row/column numbers -- if a future export adds
    or removes a patient, or shifts the rate block by a row, this still
    finds the right regions.
  - The same core transform (`parse_grid`) works whether the source arrives
    as the original .xlsx or as a delimited .txt export of the same
    layout -- `load_grid()` normalizes either into a plain 2D list first.
  - Output tables use native Excel formulas (SUMIFS / COUNTIFS / INDEX-MATCH
    / MINIFS / MAXIFS) wherever a number is a *calculation*, not raw data,
    so the workbook recalculates itself if you regenerate Sessions/Rates
    without rerunning this script. Only the melted session-level facts
    (patient, payor, session type, date, raw flag) are literal data, because
    they come from the source file, not a formula.

USAGE
-----
    python transform_attendance.py <input_path> <output_path> [--sheet NAME]

    <input_path>   .xlsx (uses the named sheet, default "Attendance and
                   Billing") or a .txt/.csv/.tsv delimited export of the
                   same grid.
    <output_path>  .xlsx to create, with tabs:
                   "Attendance and Billing" (raw, untouched copy),
                   "Rates", "Sessions", "Patients".
"""

import sys
import csv
import argparse
import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

RAW_SHEET_DEFAULT = "Attendance and Billing"
NO_SHOW_FLAG = 0.0001
ATTENDED_FLAG = 1


# ---------------------------------------------------------------------------
# Step 1: load the raw grid into a plain 2D list, regardless of source format
# ---------------------------------------------------------------------------

def load_grid(input_path: str, sheet_name: str = RAW_SHEET_DEFAULT):
    """Return (grid, source_workbook_or_None).

    grid is a list of rows, each row a list of cell values (1-indexed
    semantics preserved by leaving grid[0] as row 1, grid[0][0] as A1, etc.)
    source_workbook_or_None is the openpyxl Workbook if the input was
    .xlsx (so we can copy the raw sheet's formatting into the output
    untouched), or None if the input was a delimited text file.
    """
    path = Path(input_path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet_name]
        grid = [[cell.value for cell in row] for row in ws.iter_rows()]
        return grid, wb

    # Delimited text (.txt/.csv/.tsv) -- sniff the delimiter, parse dates.
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters="\t,")
        except csv.Error:
            dialect = csv.excel_tab
        reader = csv.reader(f, dialect)
        raw_rows = list(reader)

    def coerce(val):
        if val == "" or val is None:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.datetime.strptime(val, fmt)
            except ValueError:
                continue
        for numtype in (int, float):
            try:
                return numtype(val)
            except ValueError:
                continue
        return val

    grid = [[coerce(v) for v in row] for row in raw_rows]
    return grid, None


# ---------------------------------------------------------------------------
# Step 2: locate the rate block, date header row, and patient blocks by label
# ---------------------------------------------------------------------------

def parse_grid(grid):
    """Parse the raw wide grid into (rates, sessions, patients) records."""

    # --- Rates block: rows below the cell literally containing "Rates" ---
    rates = []  # list of dicts: session_type, payor, rate
    rates_row_idx = None
    for r, row in enumerate(grid):
        if row and isinstance(row[0], str) and row[0].strip().lower() == "rates":
            rates_row_idx = r
            break
    if rates_row_idx is None:
        raise ValueError('Could not find a "Rates" label in column A.')

    r = rates_row_idx + 1
    while r < len(grid) and grid[r] and grid[r][0]:
        label = str(grid[r][0]).strip()
        rate_val = grid[r][1]
        if rate_val in (None, ""):
            break
        # Label format observed: "IOP Commerical" / "OPT Medicaid" (note:
        # source data has a typo, "Commerical" -- normalized here).
        parts = label.replace("Commerical", "Commercial").split()
        session_type, payor = parts[0], parts[1]
        rates.append({
            "session_type": session_type,
            "payor": payor,
            "rate": float(rate_val),
        })
        r += 1
    if not rates:
        raise ValueError("Found the Rates label but no rate rows beneath it.")

    # --- Date header row: the row where column A == "Patient info" ---
    header_row_idx = None
    for r, row in enumerate(grid):
        if row and isinstance(row[0], str) and row[0].strip() == "Patient info":
            header_row_idx = r
            break
    if header_row_idx is None:
        raise ValueError('Could not find the "Patient info" header row.')

    date_row = grid[header_row_idx]
    # Dates start at column C (index 2) and run to the last non-blank col.
    date_cols = []
    for c in range(2, len(date_row)):
        v = date_row[c]
        if isinstance(v, datetime.datetime):
            date_cols.append((c, v.date()))
        elif v not in (None, ""):
            # Some sources may hand back date-as-string; try to parse.
            try:
                date_cols.append((c, datetime.datetime.fromisoformat(str(v)).date()))
            except ValueError:
                pass
    if not date_cols:
        raise ValueError("No date columns found on the header row.")

    # --- Patient blocks: pairs of rows starting right after the header ---
    sessions = []
    patients = {}  # patient_id -> payor
    r = header_row_idx + 1
    # Skip a possible second (duplicate/day-of-week) header row.
    while r < len(grid) and (not grid[r][0] or not str(grid[r][0]).startswith("Patient")):
        r += 1

    while r + 1 < len(grid):
        row_iop = grid[r]
        row_opt = grid[r + 1]
        patient_id = row_iop[0]
        payor = row_opt[0]
        type_iop = row_iop[1]
        type_opt = row_opt[1]
        if not patient_id or not str(patient_id).startswith("Patient"):
            break  # ran past the last patient block
        if type_iop != "IOP" or type_opt != "OPT":
            raise ValueError(
                f"Unexpected block shape at row {r + 1}: "
                f"expected IOP/OPT, got {type_iop!r}/{type_opt!r}"
            )
        patients[patient_id] = payor

        # Date-major, session-type-minor: same-day IOP/OPT sessions land on
        # adjacent rows (e.g. a patient's Oct 7 IOP and Oct 7 OPT sessions
        # sit next to each other) instead of being split into a whole-IOP
        # block followed by a whole-OPT block, which made same-date sessions
        # hard to spot side by side.
        for c, the_date in date_cols:
            for row_vals, session_type in ((row_iop, "IOP"), (row_opt, "OPT")):
                val = row_vals[c] if c < len(row_vals) else None
                if val in (None, ""):
                    continue  # not scheduled that day
                attended = 1 if val == ATTENDED_FLAG else 0
                sessions.append({
                    "patient_id": patient_id,
                    "payor": payor,
                    "session_type": session_type,
                    "date": the_date,
                    "raw_flag": val,
                    "attended": attended,
                })
        r += 2

    patients_out = [{"patient_id": pid, "payor": payor} for pid, payor in patients.items()]
    return rates, sessions, patients_out


# ---------------------------------------------------------------------------
# Step 3: write the output workbook (raw tab preserved + 3 new tabs)
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, name="Arial")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
BODY_FONT = Font(name="Arial")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def write_workbook(output_path, rates, sessions, patients, source_wb, sheet_name):
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    # --- Raw tab: untouched copy of the source sheet, if we have one ---
    if source_wb is not None:
        src_ws = source_wb[sheet_name]
        raw_ws = out_wb.create_sheet(sheet_name)
        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = raw_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = cell.alignment.copy()
        for col, dim in src_ws.column_dimensions.items():
            raw_ws.column_dimensions[col].width = dim.width
    # (If the source was a .txt file, there's no formatted raw tab to copy --
    # the Sessions/Patients/Rates tables below are still built the same way.)

    # --- Rates tab ---
    rates_ws = out_wb.create_sheet("Rates")
    rates_ws.append(["session_type", "payor", "rate"])
    for rt in rates:
        rates_ws.append([rt["session_type"], rt["payor"], rt["rate"]])
    for row in rates_ws.iter_rows(min_row=2):
        row[2].number_format = "$#,##0"
        for cell in row:
            cell.font = BODY_FONT
    style_header(rates_ws, 3)
    rates_last_row = len(rates) + 1

    # --- Sessions tab (long/fact table: one row per scheduled session) ---
    sess_ws = out_wb.create_sheet("Sessions")
    sess_headers = ["patient_id", "payor", "session_type", "date",
                     "raw_flag", "attended", "billable", "rate", "revenue"]
    sess_ws.append(sess_headers)
    for i, s in enumerate(sessions, start=2):
        sess_ws.cell(row=i, column=1, value=s["patient_id"])
        sess_ws.cell(row=i, column=2, value=s["payor"])
        sess_ws.cell(row=i, column=3, value=s["session_type"])
        d = sess_ws.cell(row=i, column=4, value=datetime.datetime.combine(s["date"], datetime.time()))
        d.number_format = "yyyy-mm-dd"
        sess_ws.cell(row=i, column=5, value=s["raw_flag"])
        # attended: raw_flag == 1 -> 1, else (no-show, 0.0001) -> 0
        sess_ws.cell(row=i, column=6, value=f"=IF(E{i}=1,1,0)")
        # billable == attended today; kept as its own column since billing
        # rules (e.g. auth limits, timely-filing cutoffs) may diverge from
        # pure attendance later without changing this table's shape.
        sess_ws.cell(row=i, column=7, value=f"=F{i}")
        sess_ws.cell(row=i, column=8,
                     value=f"=SUMIFS(Rates!$C$2:$C${rates_last_row},"
                           f"Rates!$A$2:$A${rates_last_row},C{i},"
                           f"Rates!$B$2:$B${rates_last_row},B{i})")
        sess_ws.cell(row=i, column=9, value=f"=G{i}*H{i}")
    for row in sess_ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row):
            cell.font = BODY_FONT
        row[7].number_format = "$#,##0"
        row[8].number_format = "$#,##0"
    style_header(sess_ws, len(sess_headers))
    sess_last_row = len(sessions) + 1

    # --- Patients tab (dimension: one row per patient) ---
    # LOS -- the headline KPI -- is los_appointments (col K): number of
    # appointments ATTENDED per patient stay, not a calendar-time span
    # (which conflates episode length with scheduling gaps). los_days/
    # los_weeks (cols G/H) are measured between a patient's first and last
    # ATTENDED appointment (cols E/F, not first/last scheduled entry in
    # cols C/D, which could include no-shows) and are kept only as a
    # secondary/internal figure -- other models in this workbook need a
    # calendar-time discharge lag, not the appointment count itself.
    # attendance_rate (col L) is each patient's own attended/scheduled
    # ratio -- already individual, not pooled.
    pat_ws = out_wb.create_sheet("Patients")
    pat_headers = ["patient_id", "payor", "first_session_date", "last_session_date",
                   "first_attended_date", "last_attended_date",
                   "los_days", "los_weeks", "los_appointments",
                   "scheduled_sessions", "attended_sessions",
                   "attendance_rate", "iop_attended", "opt_attended", "total_revenue"]
    pat_ws.append(pat_headers)
    for i, p in enumerate(patients, start=2):
        pat_ws.cell(row=i, column=1, value=p["patient_id"])
        pat_ws.cell(row=i, column=2,
                     value=f"=INDEX(Sessions!$B$2:$B${sess_last_row},"
                           f"MATCH(A{i},Sessions!$A$2:$A${sess_last_row},0))")
        c_cell = pat_ws.cell(row=i, column=3,
                     value=f"=_xlfn.MINIFS(Sessions!$D$2:$D${sess_last_row},"
                           f"Sessions!$A$2:$A${sess_last_row},A{i})")
        c_cell.number_format = "yyyy-mm-dd"
        d_cell = pat_ws.cell(row=i, column=4,
                     value=f"=_xlfn.MAXIFS(Sessions!$D$2:$D${sess_last_row},"
                           f"Sessions!$A$2:$A${sess_last_row},A{i})")
        d_cell.number_format = "yyyy-mm-dd"
        e_cell = pat_ws.cell(row=i, column=5,
                     value=f"=_xlfn.MINIFS(Sessions!$D$2:$D${sess_last_row},"
                           f"Sessions!$A$2:$A${sess_last_row},A{i},"
                           f"Sessions!$F$2:$F${sess_last_row},1)")
        e_cell.number_format = "yyyy-mm-dd"
        f_cell = pat_ws.cell(row=i, column=6,
                     value=f"=_xlfn.MAXIFS(Sessions!$D$2:$D${sess_last_row},"
                           f"Sessions!$A$2:$A${sess_last_row},A{i},"
                           f"Sessions!$F$2:$F${sess_last_row},1)")
        f_cell.number_format = "yyyy-mm-dd"
        pat_ws.cell(row=i, column=7, value=f"=F{i}-E{i}")
        pat_ws.cell(row=i, column=8, value=f"=G{i}/7")
        # los_appointments -- the headline LOS KPI: number of appointments
        # ATTENDED per patient stay (same underlying count as
        # attended_sessions, column K, surfaced here under its KPI name
        # right next to the secondary duration figures for discoverability).
        pat_ws.cell(row=i, column=9,
                     value=f"=SUMIFS(Sessions!$F$2:$F${sess_last_row},"
                           f"Sessions!$A$2:$A${sess_last_row},A{i})")
        pat_ws.cell(row=i, column=10,
                     value=f"=COUNTIFS(Sessions!$A$2:$A${sess_last_row},A{i})")
        pat_ws.cell(row=i, column=11,
                     value=f"=SUMIFS(Sessions!$F$2:$F${sess_last_row},"
                           f"Sessions!$A$2:$A${sess_last_row},A{i})")
        pat_ws.cell(row=i, column=12, value=f"=IFERROR(K{i}/J{i},0)")
        pat_ws.cell(row=i, column=12).number_format = "0.0%"
        pat_ws.cell(row=i, column=13,
                     value=f"=SUMIFS(Sessions!$F$2:$F${sess_last_row},"
                           f"Sessions!$A$2:$A${sess_last_row},A{i},"
                           f"Sessions!$C$2:$C${sess_last_row},\"IOP\")")
        pat_ws.cell(row=i, column=14,
                     value=f"=SUMIFS(Sessions!$F$2:$F${sess_last_row},"
                           f"Sessions!$A$2:$A${sess_last_row},A{i},"
                           f"Sessions!$C$2:$C${sess_last_row},\"OPT\")")
        rev_cell = pat_ws.cell(row=i, column=15,
                     value=f"=SUMIFS(Sessions!$I$2:$I${sess_last_row},"
                           f"Sessions!$A$2:$A${sess_last_row},A{i})")
        rev_cell.number_format = "$#,##0"
    for row in pat_ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
    style_header(pat_ws, len(pat_headers))

    # Column widths for readability across all new tabs.
    for ws, widths in (
        (rates_ws, [14, 12, 10]),
        (sess_ws, [14, 12, 14, 12, 10, 10, 10, 10, 12]),
        (pat_ws, [14, 12, 18, 18, 18, 18, 10, 10, 15, 18, 17, 15, 12, 12, 13]),
    ):
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    out_wb.save(output_path)


# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_path")
    parser.add_argument("output_path")
    parser.add_argument("--sheet", default=RAW_SHEET_DEFAULT)
    args = parser.parse_args()

    grid, source_wb = load_grid(args.input_path, args.sheet)
    rates, sessions, patients = parse_grid(grid)
    write_workbook(args.output_path, rates, sessions, patients, source_wb, args.sheet)

    print(f"Parsed {len(patients)} patients, {len(sessions)} scheduled sessions, "
          f"{len(rates)} rate rows.")
    print(f"Wrote {args.output_path}")


if __name__ == "__main__":
    main()
