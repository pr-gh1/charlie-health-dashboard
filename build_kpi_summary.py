"""
build_kpi_summary.py

Adds a "KPI Summary" tab to the Sessions/Patients/Rates deliverable
workbook, covering the case study's Part 1 ask directly and reviewably in
Excel: weekly LOS/census/attendance/revenue KPIs over time, LOS by
discharge cohort, and payor composition. Built off the same kpis.py
functions the dashboard uses, so the numbers match exactly.
"""
import sys
sys.path.insert(0, ".")
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill
from transform_attendance import load_grid, parse_grid
from kpis import build_frames, weekly_rollup, los_by_discharge_month, payor_composition

SRC = "sample_data/sample_attendance_export.xlsx"
DELIVERABLE_SRC = "../pipeline/sessions_output.xlsx"
OUT = "../Charlie_Health_KPI_Summary_and_Sessions.xlsx"

HEADER_FONT = Font(bold=True, name="Arial")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SECTION_FONT = Font(bold=True, name="Arial", size=13)
BODY_FONT = Font(name="Arial")

grid, wb_src = load_grid(SRC)
rates, sessions, patients = parse_grid(grid)
rates_df, sessions_df, patients_df = build_frames(rates, sessions, patients)

weekly = weekly_rollup(sessions_df, patients_df)
los = los_by_discharge_month(sessions_df, patients_df)
payor = payor_composition(sessions_df, patients_df)

shutil.copy(DELIVERABLE_SRC, OUT)
out_wb = openpyxl.load_workbook(OUT)
ws = out_wb.create_sheet("KPI Summary", 0)  # first tab, easiest to find

r = 1
ws.cell(row=r, column=1, value="Charlie Health -- Treatment & Financial KPIs Over Time").font = SECTION_FONT
r += 1
ws.cell(row=r, column=1, value="Source: Attendance and Billing tab, Aug 2020-Mar 2021. Computed via transform_attendance.py + kpis.py (same logic powering the live dashboard).").font = Font(italic=True, size=9, color="595959")
r += 2

# --- Weekly KPI table ---
ws.cell(row=r, column=1, value="Weekly KPIs").font = Font(bold=True, size=12)
r += 1
weekly_cols = [
    ("week", "Week starting", "yyyy-mm-dd"),
    ("patients_in_treatment", "Patients in treatment", "0"),
    ("iop_patients", "Patients attending IOP", "0"),
    ("new_admissions", "New admissions", "0"),
    ("attended_sessions", "Sessions attended", "0"),
    ("iop_attended", "IOP sessions billable", "0"),
    ("attendance_rate", "Attendance rate", "0.0%"),
    ("iop_attendance_rate", "IOP attendance rate", "0.0%"),
    ("avg_billed_iop_rate", "Avg billed IOP rate", "$#,##0"),
    ("avg_daily_revenue", "Avg daily revenue", "$#,##0"),
    ("revenue", "Weekly revenue", "$#,##0"),
]
header_row = r
for c, (_, label, _) in enumerate(weekly_cols, start=1):
    cell = ws.cell(row=header_row, column=c, value=label)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
r += 1
for _, row in weekly.iterrows():
    for c, (col, _, fmt) in enumerate(weekly_cols, start=1):
        val = row[col]
        cell = ws.cell(row=r, column=c, value=val.to_pydatetime() if col == "week" else float(val))
        cell.number_format = fmt
        cell.font = BODY_FONT
    r += 1
r += 2

# --- LOS by discharge cohort ---
ws.cell(row=r, column=1, value="Length of stay by discharge cohort (censoring-corrected)").font = Font(bold=True, size=12)
r += 1
ws.cell(row=r, column=1, value="LOS = appointments attended per patient stay, the headline KPI (not a calendar-time span).").font = Font(italic=True, size=9, color="595959")
r += 1
los_header = r
for c, label in enumerate(["Discharge month", "Avg LOS (appointments)", "Patients discharged"], start=1):
    cell = ws.cell(row=los_header, column=c, value=label)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
r += 1
for _, row in los.iterrows():
    ws.cell(row=r, column=1, value=str(row["discharge_month"])[:7]).font = BODY_FONT
    c2 = ws.cell(row=r, column=2, value=float(row["avg_los_appointments"])); c2.font = BODY_FONT; c2.number_format = "0.0"
    c3 = ws.cell(row=r, column=3, value=int(row["patients"])); c3.font = BODY_FONT
    r += 1
r += 2

# --- Payor composition ---
ws.cell(row=r, column=1, value="Payor composition").font = Font(bold=True, size=12)
r += 1
payor_header = r
for c, label in enumerate(["Payor", "Patients", "Avg LOS (appointments)", "Attendance rate", "Total revenue", "Avg billed IOP rate"], start=1):
    cell = ws.cell(row=payor_header, column=c, value=label)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
r += 1
for _, row in payor.iterrows():
    ws.cell(row=r, column=1, value=row["payor"]).font = BODY_FONT
    ws.cell(row=r, column=2, value=int(row["patients"])).font = BODY_FONT
    c3 = ws.cell(row=r, column=3, value=float(row["avg_los_appointments"])); c3.font = BODY_FONT; c3.number_format = "0.0"
    c4 = ws.cell(row=r, column=4, value=float(row["attendance_rate"])); c4.font = BODY_FONT; c4.number_format = "0.0%"
    c5 = ws.cell(row=r, column=5, value=float(row["total_revenue"])); c5.font = BODY_FONT; c5.number_format = "$#,##0"
    c6 = ws.cell(row=r, column=6, value=float(row["avg_billed_iop_rate"])); c6.font = BODY_FONT; c6.number_format = "$#,##0"
    r += 1

for i, w in enumerate([14, 22, 20, 16, 18, 16, 16, 16, 16, 16, 14], start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

out_wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Weekly rows: {len(weekly)}, LOS cohort rows: {len(los)}, payor rows: {len(payor)}")
